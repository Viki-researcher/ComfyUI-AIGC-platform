"""
RAG (Retrieval-Augmented Generation) Service

支持文件上传 → 文本提取 → 分块 → 嵌入 → 检索 完整 Pipeline。
支持 txt / md / pdf / docx / csv / xlsx / xls / 图片 文件类型。
"""

from __future__ import annotations

import base64
import csv
import json
import os
from pathlib import Path

import numpy as np

from app.log import logger
from app.models.chat import ChatDocument, DocumentChunk
from app.settings.config import settings


# ─── 文本提取 ───────────────────────────────────────────

_CODE_EXTENSIONS = {
    "py", "js", "ts", "jsx", "tsx", "java", "cpp", "c", "h", "hpp",
    "go", "rs", "rb", "php", "sh", "bash", "zsh",
    "yaml", "yml", "json", "xml", "html", "css", "scss", "less", "vue",
    "sql", "r", "scala", "kt", "swift", "dart", "lua",
    "toml", "ini", "cfg", "conf", "env", "dockerfile",
    "makefile", "cmake", "gradle", "pom",
}


def extract_text_from_file(file_path: str, file_type: str) -> str:
    """从文件提取纯文本。"""
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在: {file_path}")

    ext = file_type.lower()

    if ext in ("txt", "md", "markdown", "text") or ext in _CODE_EXTENSIONS:
        return path.read_text(encoding="utf-8", errors="replace")

    if ext == "pdf":
        return _extract_pdf(path)

    if ext in ("docx", "doc"):
        return _extract_docx(path)

    if ext == "csv":
        return _extract_csv(path)

    if ext in ("xlsx", "xls"):
        return _extract_xlsx(path)

    if ext in ("png", "jpg", "jpeg", "gif", "webp"):
        return f"Image file: {path.name}"

    return path.read_text(encoding="utf-8", errors="replace")


def _extract_pdf(path: Path) -> str:
    from pypdf import PdfReader
    reader = PdfReader(str(path))
    parts = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            parts.append(text)
    return "\n".join(parts)


def _extract_docx(path: Path) -> str:
    from docx import Document
    doc = Document(str(path))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _extract_csv(path: Path) -> str:
    """读取 CSV 文件并格式化为 Markdown 表格。"""
    with path.open(encoding="utf-8", errors="replace", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        return ""

    return _rows_to_markdown_table(rows)


def _extract_xlsx(path: Path) -> str:
    """读取 xlsx/xls 文件第一个 sheet 并格式化为 Markdown 表格。"""
    from openpyxl import load_workbook
    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return ""

    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(cell) if cell is not None else "" for cell in row])
    wb.close()

    if not rows:
        return ""

    return _rows_to_markdown_table(rows)


def _rows_to_markdown_table(rows: list[list[str]]) -> str:
    """将行列表转换为 Markdown 表格字符串。"""
    if not rows:
        return ""

    header = rows[0]
    col_count = len(header)

    lines = []
    lines.append("| " + " | ".join(str(h) for h in header) + " |")
    lines.append("| " + " | ".join("---" for _ in range(col_count)) + " |")

    for row in rows[1:]:
        padded = list(row) + [""] * (col_count - len(row))
        lines.append("| " + " | ".join(str(c) for c in padded[:col_count]) + " |")

    return "\n".join(lines)


# ─── 图片描述（多模态 RAG） ─────────────────────────────

async def extract_image_description(file_path: str) -> str:
    """使用 LLM 视觉能力描述图片内容，失败时回退到文件名。"""
    path = Path(file_path)
    filename = path.name

    try:
        from app.services.llm_client import chat_completion

        image_data = path.read_bytes()
        b64 = base64.b64encode(image_data).decode("utf-8")

        suffix = path.suffix.lower().lstrip(".")
        mime_map = {
            "png": "image/png",
            "jpg": "image/jpeg",
            "jpeg": "image/jpeg",
            "gif": "image/gif",
            "webp": "image/webp",
        }
        mime = mime_map.get(suffix, "image/png")

        messages = [
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": "请详细描述这张图片的内容，包括主要元素、文字、颜色和布局。",
                    },
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:{mime};base64,{b64}",
                        },
                    },
                ],
            }
        ]

        description = await chat_completion(messages)
        if description and description.strip():
            return description.strip()
    except Exception as e:
        logger.warning(f"[RAG] image description failed for {filename}: {e}")

    return f"Image file: {filename}"


# ─── 文本分块 ───────────────────────────────────────────

def chunk_text(text: str, chunk_size: int = 0, overlap: int = 0) -> list[str]:
    """将文本按字符长度切分为多个块，支持重叠。"""
    if not chunk_size:
        chunk_size = settings.RAG_CHUNK_SIZE
    if not overlap:
        overlap = settings.RAG_CHUNK_OVERLAP

    text = text.strip()
    if not text:
        return []

    if len(text) <= chunk_size:
        return [text]

    chunks: list[str] = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunks.append(text[start:end])
        start = end - overlap

    return chunks


# ─── 嵌入生成 & 存储 ──────────────────────────────────────

async def process_document(doc_id: int) -> None:
    """完整处理上传文档: 提取 → 分块 → 嵌入 → 存储。"""
    doc = await ChatDocument.get_or_none(id=doc_id)
    if not doc:
        return

    try:
        ext = doc.file_type.lower()

        # 图片文件使用视觉模型描述
        if ext in ("png", "jpg", "jpeg", "gif", "webp"):
            text = await extract_image_description(doc.file_path)
        else:
            text = extract_text_from_file(doc.file_path, doc.file_type)

        if not text.strip():
            doc.status = "error"
            doc.error_msg = "文件内容为空"
            await doc.save()
            return

        chunks = chunk_text(text)
        logger.info(f"[RAG] doc={doc_id} extracted {len(chunks)} chunks")

        for idx, chunk_text_content in enumerate(chunks):
            embedding_json = ""
            try:
                from app.services.llm_client import generate_embedding
                emb = await generate_embedding(chunk_text_content[:2000])
                embedding_json = json.dumps(emb)
            except Exception as e:
                logger.warning(f"[RAG] embedding failed for chunk {idx}: {e}")

            await DocumentChunk.create(
                document_id=doc.id,
                content=chunk_text_content,
                chunk_index=idx,
                embedding=embedding_json,
            )

        doc.chunk_count = len(chunks)
        doc.status = "ready"
        await doc.save()
        logger.info(f"[RAG] doc={doc_id} processing complete, {len(chunks)} chunks stored")

    except Exception as e:
        logger.error(f"[RAG] doc={doc_id} processing failed: {e}")
        doc.status = "error"
        doc.error_msg = str(e)[:500]
        await doc.save()


# ─── 检索 ─────────────────────────────────────────────

def _cosine_similarity(a: list[float], b: list[float]) -> float:
    va = np.array(a, dtype=np.float32)
    vb = np.array(b, dtype=np.float32)
    dot = float(np.dot(va, vb))
    norm = float(np.linalg.norm(va) * np.linalg.norm(vb))
    return dot / norm if norm > 0 else 0.0


async def retrieve_relevant_chunks(
    query: str,
    user_id: int,
    document_ids: list[int] | None = None,
    top_k: int = 0,
) -> list[dict]:
    """根据查询文本检索最相关的文档分块，返回含文档名的结果。"""
    if not top_k:
        top_k = settings.RAG_TOP_K

    # 生成查询嵌入
    try:
        from app.services.llm_client import generate_embedding
        query_emb = await generate_embedding(query[:2000])
    except Exception as e:
        logger.warning(f"[RAG] query embedding failed, falling back to keyword search: {e}")
        return await _keyword_search(query, user_id, document_ids, top_k)

    # 获取用户的所有文档块
    doc_filter = {"user_id": user_id, "status": "ready"}
    docs = await ChatDocument.filter(**doc_filter).all()
    if document_ids:
        docs = [d for d in docs if d.id in document_ids]

    if not docs:
        return []

    doc_name_map = {d.id: d.filename for d in docs}
    doc_ids = list(doc_name_map.keys())

    chunks = await DocumentChunk.filter(document_id__in=doc_ids).all()

    # 向量相似度排序
    scored: list[tuple[float, DocumentChunk]] = []
    for chunk in chunks:
        if not chunk.embedding:
            continue
        try:
            emb = json.loads(chunk.embedding)
            score = _cosine_similarity(query_emb, emb)
            scored.append((score, chunk))
        except (json.JSONDecodeError, ValueError):
            continue

    scored.sort(key=lambda x: x[0], reverse=True)

    results = []
    for score, chunk in scored[:top_k]:
        results.append({
            "chunk_id": chunk.id,
            "document_id": chunk.document_id,
            "document_name": doc_name_map.get(chunk.document_id, ""),
            "content": chunk.content,
            "chunk_index": chunk.chunk_index,
            "score": round(score, 4),
        })
    return results


async def _keyword_search(
    query: str,
    user_id: int,
    document_ids: list[int] | None,
    top_k: int,
) -> list[dict]:
    """关键词回退搜索（当嵌入不可用时）。支持中文文本的模糊匹配。"""
    doc_filter = {"user_id": user_id, "status": "ready"}
    docs = await ChatDocument.filter(**doc_filter).all()
    if document_ids:
        docs = [d for d in docs if d.id in document_ids]
    if not docs:
        return []

    doc_name_map = {d.id: d.filename for d in docs}
    doc_ids = list(doc_name_map.keys())

    chunks = await DocumentChunk.filter(document_id__in=doc_ids).all()
    import re

    raw_keywords = [w for w in re.split(r'[\s,，。？！?!、；;：:·\-—\u201c\u201d\u2018\u2019()（）【】\[\]]+', query.lower()) if len(w) >= 2]
    keywords = set(raw_keywords)
    for kw in raw_keywords:
        if len(kw) >= 4:
            for i in range(len(kw) - 1):
                keywords.add(kw[i:i + 2])
    keywords = list(keywords) or [query.lower().strip()[:6]]

    scored = []
    for chunk in chunks:
        content_lower = chunk.content.lower()
        score = sum(1 for kw in keywords if kw in content_lower)
        if score > 0:
            scored.append((score, chunk))

    if not scored and chunks:
        scored = [(1, chunks[0])]

    scored.sort(key=lambda x: x[0], reverse=True)
    return [
        {
            "chunk_id": c.id,
            "document_id": c.document_id,
            "document_name": doc_name_map.get(c.document_id, ""),
            "content": c.content,
            "chunk_index": c.chunk_index,
            "score": s,
        }
        for s, c in scored[:top_k]
    ]


# ─── ComfyUI 文档索引 ──────────────────────────────────

async def index_comfyui_docs(comfyui_repo_path: str) -> int:
    """扫描 ComfyUI 仓库的文档和代码文件，创建索引。

    只处理尚未索引的文件（按 file_path 去重）。
    所有条目使用 user_id=0 (系统文档)。
    返回新创建的 chunk 数量。
    """
    repo = Path(comfyui_repo_path)
    if not repo.is_dir():
        logger.error(f"[RAG] ComfyUI repo not found: {comfyui_repo_path}")
        return 0

    target_extensions = {".md", ".txt", ".py"}
    skip_dirs = {".git", "__pycache__", "node_modules", ".venv", "venv"}

    files_to_index: list[Path] = []
    for root, dirs, filenames in os.walk(repo):
        dirs[:] = [d for d in dirs if d not in skip_dirs]
        for fname in filenames:
            fpath = Path(root) / fname
            if fpath.suffix.lower() in target_extensions:
                files_to_index.append(fpath)

    if not files_to_index:
        logger.info("[RAG] No indexable files found in ComfyUI repo")
        return 0

    # 获取已索引的文件路径集合
    existing_docs = await ChatDocument.filter(user_id=0).values_list("file_path", flat=True)
    existing_paths = set(existing_docs)

    total_chunks = 0

    for fpath in files_to_index:
        fpath_str = str(fpath)
        if fpath_str in existing_paths:
            continue

        try:
            ext = fpath.suffix.lower().lstrip(".")
            text = extract_text_from_file(fpath_str, ext)
            if not text.strip():
                continue

            doc = await ChatDocument.create(
                user_id=0,
                filename=fpath.name,
                file_path=fpath_str,
                file_size=fpath.stat().st_size,
                file_type=ext,
                status="processing",
            )

            chunks = chunk_text(text)
            for idx, chunk_content in enumerate(chunks):
                embedding_json = ""
                try:
                    from app.services.llm_client import generate_embedding
                    emb = await generate_embedding(chunk_content[:2000])
                    embedding_json = json.dumps(emb)
                except Exception as e:
                    logger.warning(f"[RAG] embedding failed for {fpath.name} chunk {idx}: {e}")

                await DocumentChunk.create(
                    document_id=doc.id,
                    content=chunk_content,
                    chunk_index=idx,
                    embedding=embedding_json,
                )

            doc.chunk_count = len(chunks)
            doc.status = "ready"
            await doc.save()
            total_chunks += len(chunks)

        except Exception as e:
            logger.warning(f"[RAG] failed to index {fpath}: {e}")
            continue

    logger.info(f"[RAG] ComfyUI docs indexing complete: {total_chunks} new chunks")
    return total_chunks


def build_rag_context(chunks: list[dict]) -> str:
    """将检索到的分块构建为带编号来源标注的上下文文本，引导 LLM 在回答中引用来源。"""
    if not chunks:
        return ""
    parts = [
        "以下是从用户上传的文档中检索到的相关片段。请基于这些内容回答用户问题。"
        "回答时**必须**在相关语句后标注引用来源，格式为 [来源1]、[来源2] 等。"
        "如果某段内容未被使用，不要引用。引用时请指出具体是哪个文档的哪部分。",
        "",
    ]
    for i, ch in enumerate(chunks, 1):
        doc_name = ch.get("document_name", "")
        chunk_index = ch.get("chunk_index", "")
        parts.append(f"[来源{i}] 文档「{doc_name}」第 {chunk_index} 片段:")
        parts.append(ch["content"])
        parts.append("")
    return "\n".join(parts)


def ensure_upload_dir() -> str:
    upload_dir = settings.RAG_UPLOAD_DIR
    os.makedirs(upload_dir, exist_ok=True)
    return upload_dir
